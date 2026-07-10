from io import BytesIO
from re import sub
from urllib.parse import quote
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from tortoise.transactions import in_transaction

from app.api.v1.crud_router_factory import serialize_model
from app.core.auth.dependencies import require_komite_employee
from app.models.entities import (
    Condominium,
    CondominiumInspectionItem,
    CondominiumInspectionTemplate,
    InspectionTemplate,
    InspectionTemplateItem,
)
from app.schemas.entity_schemas import (
    CondominiumInspectionTemplateOut,
    InspectionTemplateDuplicateToCondominiumOut,
    InspectionTemplateDuplicateToCondominiumRequest,
)


PERIODICITY_LABELS = {
    "daily": "Diaria",
    "weekly": "Semanal",
    "biweekly": "Quincenal",
    "monthly": "Mensual",
    "bimonthly": "Cada 2 meses",
    "quarterly": "Trimestral",
    "four_monthly": "Cada 4 meses",
    "semiannual": "Semestral",
    "annual": "Anual",
    "biennial": "Cada 2 años",
    "permanent": "Permanente",
    "on_demand": "Según necesidad",
}

MONTH_LABELS = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


router = APIRouter(
    prefix="/api/v1/inspection-templates",
    tags=["Inspection templates"],
    dependencies=[Depends(require_komite_employee())],
)


def _actor(request: Request) -> str:
    user = getattr(request.state, "user", None)
    return getattr(user, "email", None) or getattr(request.state, "user_id", None) or "backoffice"


def _safe_filename(value: str) -> str:
    clean = sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return clean.strip("_") or "plantilla"


def _format_months(value) -> str:
    if not value:
        return ""
    labels = []
    for item in value:
        try:
            labels.append(MONTH_LABELS.get(int(item), str(item)))
        except (TypeError, ValueError):
            labels.append(str(item))
    return ", ".join(labels)


def _autosize_columns(sheet, max_width: int = 52) -> None:
    for column_cells in sheet.columns:
        column = get_column_letter(column_cells[0].column)
        width = min(
            max_width,
            max(len(str(cell.value or "")) for cell in column_cells) + 2,
        )
        sheet.column_dimensions[column].width = max(width, 12)


def _style_table(sheet, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="102437")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = f"A{header_row + 1}"
    _autosize_columns(sheet)


async def _get_template_or_404(template_id: UUID) -> InspectionTemplate:
    template = await InspectionTemplate.get_or_none(id=template_id)
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Plantilla base no encontrada")
    return template


@router.get("/{template_id}/export")
async def export_template(template_id: UUID):
    template = await _get_template_or_404(template_id)
    sections = await template.sections.all().order_by("display_order", "name")
    items = await InspectionTemplateItem.filter(template_id=template_id).select_related("section").order_by(
        "display_order",
        "task_name",
    )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    summary.append(["Campo", "Valor"])
    summary.append(["Plantilla", template.name])
    summary.append(["Descripción", template.description or ""])
    summary.append(["Tipo de plantilla", template.template_type])
    summary.append(["Categoría", template.inspection_type])
    summary.append(["Versión", template.version])
    summary.append(["Estado", template.status])
    summary.append(["Activa", "Sí" if template.is_active else "No"])
    summary.append(["Archivo origen", template.source_file_name or ""])
    summary.append(["Total secciones", len(sections)])
    summary.append(["Total ítems", len(items)])
    _style_table(summary)

    section_sheet = workbook.create_sheet("Secciones")
    section_sheet.append(["Orden", "Sección", "Descripción", "Estado"])
    for section in sections:
        section_sheet.append([
            section.display_order,
            section.name,
            section.description or "",
            section.status,
        ])
    _style_table(section_sheet)

    items_sheet = workbook.create_sheet("Items")
    items_sheet.append([
        "Orden",
        "Sección",
        "Activo o zona",
        "Tarea",
        "Instrucciones",
        "Periodicidad",
        "Meses planificados",
        "Responsable sugerido",
        "Duración estimada min.",
        "Requiere evidencia",
        "Estado",
    ])
    for item in items:
        items_sheet.append([
            item.display_order,
            item.section.name if item.section else "",
            item.asset_name or "",
            item.task_name,
            item.instructions or "",
            PERIODICITY_LABELS.get(item.periodicity, item.periodicity or ""),
            _format_months(item.planned_months),
            item.default_responsible_profile or "",
            item.default_duration_minutes or "",
            "Sí" if item.requires_evidence else "No",
            item.status,
        ])
    _style_table(items_sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"plantilla_{_safe_filename(template.name)}.xlsx"
    encoded_filename = quote(filename)
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post(
    "/{template_id}/duplicate-to-condominium",
    response_model=InspectionTemplateDuplicateToCondominiumOut,
    status_code=status.HTTP_201_CREATED,
)
async def duplicate_template_to_condominium(
    template_id: UUID,
    payload: InspectionTemplateDuplicateToCondominiumRequest,
    request: Request,
):
    base_template = await _get_template_or_404(template_id)

    condominium = await Condominium.get_or_none(id=payload.condominium_id)
    if not condominium:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Condominio no encontrado")

    base_items = await InspectionTemplateItem.filter(template_id=template_id).select_related("section").order_by(
        "display_order",
        "task_name",
    )
    if not base_items:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="La plantilla base no tiene ítems para duplicar",
        )

    last_version = (
        await CondominiumInspectionTemplate.filter(
            condominium_id=condominium.id,
            base_template_id=base_template.id,
        )
        .order_by("-version")
        .first()
    )
    next_version = (last_version.version + 1) if last_version else 1
    actor = _actor(request)
    copy_name = payload.name or f"{base_template.name} - {condominium.name}"

    async with in_transaction() as connection:
        condominium_template = await CondominiumInspectionTemplate.create(
            using_db=connection,
            company_id=condominium.company_id,
            condominium_id=condominium.id,
            base_template_id=base_template.id,
            name=copy_name,
            template_type=base_template.template_type,
            version=next_version,
            status=payload.status,
            metadata={
                "source_template_id": str(base_template.id),
                "source_template_name": base_template.name,
                "source_template_version": base_template.version,
                "duplicated_from_backoffice": True,
            },
            created_by=actor,
            updated_by=actor,
        )

        for base_item in base_items:
            await CondominiumInspectionItem.create(
                using_db=connection,
                company_id=condominium.company_id,
                condominium_id=condominium.id,
                condominium_template_id=condominium_template.id,
                base_item_id=base_item.id,
                section_name=base_item.section.name if base_item.section else None,
                asset_name=base_item.asset_name,
                task_name=base_item.task_name,
                instructions=base_item.instructions,
                event_type=base_item.event_type,
                periodicity=base_item.periodicity,
                planned_months=base_item.planned_months,
                responsible_profile=base_item.default_responsible_profile,
                estimated_duration_minutes=base_item.default_duration_minutes,
                priority="medium",
                status=base_item.status,
                metadata={
                    **(base_item.metadata or {}),
                    "source_template_item_id": str(base_item.id),
                    "source_template_id": str(base_template.id),
                },
                created_by=actor,
                updated_by=actor,
            )

    return InspectionTemplateDuplicateToCondominiumOut(
        template=CondominiumInspectionTemplateOut(**serialize_model(condominium_template)),
        items_created=len(base_items),
    )
