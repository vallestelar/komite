from __future__ import annotations

import unicodedata
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.models.entities import Unit, UnitContact, User


class ComunidadFelizNeighborsImportService:
    relationship_owner = "copropietario"
    relationship_resident = "residente"
    relationship_tenant = "arrendatario"
    relationship_contact = "contacto"

    async def import_charges(
        self,
        *,
        company_id: str | None,
        condominium_id: str,
        content: bytes,
    ) -> dict[str, Any]:
        rows = self._read_rows(content)
        summary = self._empty_summary(rows)
        details: list[dict[str, str]] = []
        processed_units: set[str] = set()

        for row in rows:
            unit, created_unit = await self._upsert_unit(company_id, condominium_id, row)
            if unit.identifier not in processed_units:
                summary["units_created" if created_unit else "units_updated"] += 1
                processed_units.add(unit.identifier)

            full_name = self._clean(row.get("resident_name"))
            if not full_name:
                summary["contacts_skipped"] += 1
                continue

            email = self._clean_email(row.get("email"))
            document_number = self._clean(row.get("document_number"))
            relationship_type = row.get("relationship_type") or self.relationship_resident
            user_status = await self._classify_user(company_id=company_id, email=email)
            contact, created_contact = await self._upsert_contact(
                company_id=company_id,
                condominium_id=condominium_id,
                unit=unit,
                relationship_type=relationship_type,
                full_name=full_name,
                email=email,
                phone=self._clean(row.get("phone")),
                document_number=document_number,
                is_primary_contact=self._bool_from_text(row.get("is_primary_contact")),
                row=row,
            )
            summary["contacts_created" if created_contact else "contacts_updated"] += 1
            summary[f"users_{user_status}"] += 1
            details.append(
                {
                    "unit": unit.identifier,
                    "relationship_type": contact.relationship_type,
                    "full_name": contact.full_name,
                    "status": "creado" if created_contact else "actualizado",
                }
            )

        return {"summary": summary, "items": details[:200]}

    async def preview_charges(
        self,
        *,
        company_id: str | None,
        condominium_id: str,
        content: bytes,
    ) -> dict[str, Any]:
        rows = self._read_rows(content)
        summary = self._empty_summary(rows)
        details: list[dict[str, str]] = []
        processed_units: set[str] = set()

        for row in rows:
            identifier = self._clean(row.get("uco")) or "Sin identificador"
            unit = await Unit.get_or_none(condominium_id=condominium_id, identifier=identifier)
            if identifier not in processed_units:
                summary["units_updated" if unit else "units_created"] += 1
                processed_units.add(identifier)

            full_name = self._clean(row.get("resident_name"))
            if not full_name:
                summary["contacts_skipped"] += 1
                continue

            email = self._clean_email(row.get("email"))
            document_number = self._clean(row.get("document_number"))
            relationship_type = row.get("relationship_type") or self.relationship_resident
            existing_contact = await self._find_contact(unit, relationship_type, email, document_number, full_name) if unit else None
            contact_status = "updated" if existing_contact else "created"
            summary[f"contacts_{contact_status}"] += 1

            user_status = await self._classify_user(company_id=company_id, email=email)
            summary[f"users_{user_status}"] += 1
            details.append(
                {
                    "unit": identifier,
                    "relationship_type": relationship_type,
                    "full_name": full_name,
                    "status": "se va a crear" if contact_status == "created" else "se va a actualizar",
                }
            )

        return {"summary": summary, "items": details[:200]}

    def _empty_summary(self, rows: list[dict[str, Any]]) -> dict[str, int]:
        return {
            "rows": len(rows),
            "format_residents": 1 if rows and rows[0].get("source_format") == "residents" else 0,
            "units_created": 0,
            "units_updated": 0,
            "contacts_created": 0,
            "contacts_updated": 0,
            "contacts_skipped": 0,
            "users_created": 0,
            "users_updated": 0,
            "users_skipped": 0,
        }

    def _read_rows(self, content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = (
            workbook["Listado de Copropietarios"]
            if "Listado de Copropietarios" in workbook.sheetnames
            else workbook["Gastos comunes"]
            if "Gastos comunes" in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        header_row = self._find_header_row(sheet)
        headers = [self._cell_text(cell.value) for cell in sheet[header_row]]
        indexes = self._build_indexes(headers)

        if indexes.get("unidad") is None:
            raise ValueError("El informe no tiene la columna Unidad esperada.")

        is_residents_format = indexes.get("nombre_apellido") is not None and indexes.get("rol") is not None
        if not is_residents_format and indexes.get("residente") is None:
            raise ValueError("El informe no tiene las columnas Residente o Nombre y apellido esperadas.")

        rows: list[dict[str, Any]] = []
        for raw in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            unit_text = self._clean(self._value(raw, indexes["unidad"]))
            if not unit_text or unit_text.casefold() == "total":
                continue

            if is_residents_format:
                role = self._cell_text(self._value(raw, indexes.get("rol")))
                rows.append(
                    {
                        "source_format": "residents",
                        "uco": unit_text,
                        "resident_name": self._cell_text(self._value(raw, indexes.get("nombre_apellido"))),
                        "relationship_type": self._map_relationship(role),
                        "source_role": role,
                        "role_status": self._cell_text(self._value(raw, indexes.get("estado_rol"))),
                        "is_primary_contact": self._value(raw, indexes.get("encargado")),
                        "email": self._cell_text(self._value(raw, indexes.get("correo"))),
                        "phone": self._cell_text(self._value(raw, indexes.get("telefono"))),
                        "document_number": self._cell_text(self._value(raw, indexes.get("rut"))),
                        "system_access": self._cell_text(self._value(raw, indexes.get("ingreso_sistema"))),
                        "last_login": self._cell_text(self._value(raw, indexes.get("ultimo_ingreso"))),
                        "proration": self._value(raw, indexes.get("prorrateo")),
                    }
                )
            else:
                rows.append(
                    {
                        "source_format": "charges",
                        "uco": unit_text,
                        "resident_name": self._cell_text(self._value(raw, indexes.get("residente"))),
                        "relationship_type": self.relationship_resident,
                        "proration": self._value(raw, indexes.get("prorrateo")),
                        "month_total": self._value(raw, indexes.get("total_mes")),
                        "past_due_capital": self._value(raw, indexes.get("capital_atrasado")),
                        "interests_and_fines": self._value(raw, indexes.get("intereses_multas")),
                        "payments": self._value(raw, indexes.get("abonos")),
                        "charged_total": self._value(raw, indexes.get("total_cobrado")),
                        "credit_balance": self._value(raw, indexes.get("saldo_favor")),
                        "source_period": self._detect_period(sheet),
                    }
                )

        if not rows:
            raise ValueError("El informe de Comunidad Feliz no contiene filas para importar.")
        return rows

    async def _upsert_unit(self, company_id: str | None, condominium_id: str, row: dict[str, Any]) -> tuple[Unit, bool]:
        identifier = self._clean(row.get("uco")) or "Sin identificador"
        unit = await Unit.get_or_none(condominium_id=condominium_id, identifier=identifier)
        created = unit is None
        if not unit:
            unit = Unit(company_id=company_id, condominium_id=condominium_id, identifier=identifier)

        unit.company_id = company_id
        unit.external_code = identifier
        unit.unit_type = "apartment"
        unit.proration_total = self._decimal_or_none(row.get("proration"))
        unit.proration = self._decimal_or_none(row.get("proration"))
        unit.metadata = {
            **(unit.metadata or {}),
            "comunidad_feliz": {
                "format": row.get("source_format"),
                "period": row.get("source_period"),
                "month_total": self._number_or_none(row.get("month_total")),
                "past_due_capital": self._number_or_none(row.get("past_due_capital")),
                "interests_and_fines": self._number_or_none(row.get("interests_and_fines")),
                "payments": self._number_or_none(row.get("payments")),
                "charged_total": self._number_or_none(row.get("charged_total")),
                "credit_balance": self._number_or_none(row.get("credit_balance")),
            },
        }
        await unit.save()
        return unit, created

    async def _upsert_contact(
        self,
        *,
        company_id: str | None,
        condominium_id: str,
        unit: Unit,
        relationship_type: str,
        full_name: str,
        email: str | None,
        phone: str | None,
        document_number: str | None,
        is_primary_contact: bool,
        row: dict[str, Any],
    ) -> tuple[UnitContact, bool]:
        contact = await self._find_contact(unit, relationship_type, email, document_number, full_name)
        created = contact is None
        user = await self._sync_user(company_id, email, full_name, phone, document_number)
        if not contact:
            contact = UnitContact(
                company_id=company_id,
                condominium_id=condominium_id,
                unit=unit,
                relationship_type=relationship_type,
                full_name=full_name,
            )

        contact.company_id = company_id
        contact.condominium_id = condominium_id
        contact.unit = unit
        contact.user = user
        contact.full_name = full_name
        contact.relationship_type = relationship_type
        contact.email = email
        contact.phone = phone
        contact.document_type = "rut" if document_number else None
        contact.document_number = document_number
        contact.is_primary_contact = is_primary_contact
        if created and not email:
            contact.receives_notifications = False
        contact.status = "active"
        contact.metadata = {
            **(contact.metadata or {}),
            "source": "comunidad_feliz",
            "source_format": row.get("source_format"),
            "source_period": row.get("source_period"),
            "source_role": row.get("source_role"),
            "role_status": row.get("role_status"),
            "system_access": row.get("system_access"),
            "last_login": row.get("last_login"),
        }
        await contact.save()
        return contact, created

    async def _find_contact(
        self,
        unit: Unit | None,
        relationship_type: str,
        email: str | None,
        document_number: str | None,
        full_name: str,
    ) -> UnitContact | None:
        if not unit:
            return None
        base = UnitContact.filter(unit_id=unit.id, relationship_type=relationship_type)
        if email:
            return await base.filter(email=email).first()
        if document_number:
            return await base.filter(document_number=document_number).first()
        return await base.filter(full_name=full_name).first()

    async def _classify_user(self, company_id: str | None, email: str | None) -> str:
        if not email:
            return "skipped"
        user = await User.get_or_none(email=email)
        if not user:
            return "created"
        if company_id and str(user.company_id) == str(company_id):
            return "updated"
        return "skipped"

    async def _sync_user(
        self,
        company_id: str | None,
        email: str | None,
        full_name: str,
        phone: str | None,
        document_number: str | None,
    ) -> User | None:
        if not email:
            return None
        user = await User.get_or_none(email=email)
        if not user:
            return await User.create(
                company_id=company_id,
                email=email,
                full_name=full_name,
                phone=phone,
                document_type="rut" if document_number else None,
                document_number=document_number,
                status="active",
            )
        if company_id and str(user.company_id) == str(company_id):
            user.full_name = full_name
            user.phone = phone
            user.document_type = "rut" if document_number else user.document_type
            user.document_number = document_number
            await user.save()
            return user
        return None

    def _find_header_row(self, sheet) -> int:
        for index, row in enumerate(sheet.iter_rows(), start=1):
            normalized = {self._normalize_header(self._cell_text(cell.value)) for cell in row}
            if {"unidad", "residente"}.issubset(normalized) or {"unidad", "nombreyapellido", "rol"}.issubset(normalized):
                return index
        raise ValueError("No se encontro la cabecera del informe de Comunidad Feliz.")

    def _build_indexes(self, headers: list[str]) -> dict[str, int | None]:
        groups: dict[str, list[int]] = {}
        for index, header in enumerate(headers):
            normalized = self._normalize_header(header)
            groups.setdefault(normalized, []).append(index)

        def idx(name: str) -> int | None:
            values = groups.get(self._normalize_header(name), [])
            return values[0] if values else None

        return {
            "unidad": idx("Unidad"),
            "residente": idx("Residente"),
            "nombre_apellido": idx("Nombre y apellido"),
            "rol": idx("Rol"),
            "estado_rol": idx("Estado de rol"),
            "encargado": idx("Encargado"),
            "correo": idx("Correo"),
            "telefono": idx("Telefono"),
            "rut": idx("RUT"),
            "ingreso_sistema": idx("Ingreso al sistema"),
            "ultimo_ingreso": idx("Ultimo ingreso"),
            "prorrateo": idx("Prorrateo"),
            "total_mes": idx("Total mes"),
            "capital_atrasado": idx("Capital Atrasado"),
            "intereses_multas": idx("Intereses y multas"),
            "abonos": idx("Abonos"),
            "total_cobrado": idx("Total cobrado"),
            "saldo_favor": idx("Saldo a favor"),
        }

    def _detect_period(self, sheet) -> str | None:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 4), values_only=True):
            for value in row:
                text = self._cell_text(value)
                if text.casefold().startswith("gastos comunes"):
                    return text
        return None

    def _value(self, row: tuple[Any, ...], index: int | None) -> Any:
        if index is None or index >= len(row):
            return None
        return row[index]

    def _cell_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _normalize_header(self, value: str) -> str:
        text = unicodedata.normalize("NFKD", value)
        return "".join(ch for ch in text.casefold() if ch.isalnum() and not unicodedata.combining(ch))

    def _clean(self, value: Any) -> str | None:
        text = self._cell_text(value)
        return text if text else None

    def _clean_email(self, value: Any) -> str | None:
        text = self._clean(value)
        return text.casefold() if text and "@" in text else None

    def _map_relationship(self, value: Any) -> str:
        normalized = self._normalize_header(self._clean(value) or "")
        if "dueno" in normalized or "propietario" in normalized:
            return self.relationship_owner
        if "arrendatario" in normalized or "arrendador" in normalized:
            return self.relationship_tenant
        if "residente" in normalized:
            return self.relationship_resident
        return self.relationship_contact

    def _bool_from_text(self, value: Any) -> bool:
        normalized = self._normalize_header(self._clean(value) or "")
        return normalized in {"si", "true", "1", "yes"}

    def _decimal_or_none(self, value: Any) -> Decimal | None:
        text = self._clean(value)
        if not text:
            return None
        try:
            return Decimal(text.replace(",", "."))
        except (InvalidOperation, ValueError):
            return None

    def _number_or_none(self, value: Any) -> int | float | None:
        if value is None or value == "":
            return None
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (int, float)):
            return value
        try:
            number = Decimal(str(value).replace(",", "."))
        except (InvalidOperation, ValueError):
            return None
        return int(number) if number == number.to_integral_value() else float(number)
