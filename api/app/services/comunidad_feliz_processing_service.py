from __future__ import annotations

import base64
import re
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.services.edifito_processing_service import EdifitoProcessingService

try:
    import xlrd
except ImportError:  # pragma: no cover - dependency is declared for runtime
    xlrd = None


@dataclass
class ChileBankMovement:
    monto: str
    nombre: str
    fecha: str
    documento: str
    sucursal: str
    saldo: str
    rut: str = ""


@dataclass
class ComunidadFelizCharge:
    uco: str
    rut: str
    nombre: str
    cobro: str
    pago: str
    estado: str
    fecha_pago: str


class ComunidadFelizProcessingService:
    output_columns = [
        ("monto", "Monto"),
        ("rut", "RUT"),
        ("nombre", "Nombre"),
        ("fecha", "Fecha"),
        ("documento", "Documento"),
        ("sucursal", "Sucursal"),
        ("uco", "UCO"),
        ("match_rut", "MatchRUT"),
        ("match_nombre", "MatchNombre"),
        ("mes", "Mes"),
        ("cobro", "Cobro"),
        ("pago", "Pago"),
        ("fecha_pago", "FechaPago"),
        ("status", "Status"),
    ]

    def process(self, *, bank_name: str, bank_statement: bytes, bank_filename: str, charges_file: bytes) -> dict[str, Any]:
        normalized_bank = (bank_name or "").casefold()
        if "santander" in normalized_bank:
            movements = self._read_santander(bank_statement)
        elif "scotia" in normalized_bank:
            movements = self._read_scotiabank(bank_statement, bank_filename)
        elif "chile" in normalized_bank:
            movements = self._read_bank_chile(bank_statement, bank_filename)
        else:
            raise ValueError("Comunidad Feliz procesa cartolas Santander, Banco de Chile o Scotiabank.")

        charges = self._read_charges(charges_file)
        rows = self._build_rows(movements, charges)

        return {
            "rows": rows,
            "summary": self._summary(rows),
            "download_filename": f"Movimientos_Comunidad_Feliz_{datetime.now().strftime('%d_%m_%Y_%H%M')}.xlsx",
            "download_base64": self.build_movements_workbook_base64(rows),
            "ingresos_filename": f"ingresos_comunidad_feliz_{datetime.now().strftime('%d_%m_%Y_%H%M')}.xlsx",
            "ingresos_base64": self.build_ingresos_workbook_base64(rows),
        }

    def build_movements_workbook_base64(self, rows: list[dict[str, str]]) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Movimientos"
        sheet.append([label for _, label in self.output_columns])

        header_fill = PatternFill("solid", fgColor="D9E2F3")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row in rows:
            sheet.append([row.get(key, "") for key, _ in self.output_columns])

        self._fit_columns(sheet)
        return self._workbook_to_base64(workbook)

    def build_ingresos_workbook_base64(self, rows: list[dict[str, str]]) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Ingresos"
        headers = [
            "Folio",
            "Propiedad",
            "Monto a pagar",
            "Fecha de pago",
            "Medio de pago",
            "Numero documento de pago",
            "Comentarios del pago",
        ]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E2F0D9")

        folio = 1
        for row in rows:
            if row.get("status") != "Procesar Pago":
                continue
            amount = self._amount_int(row.get("monto", ""))
            charge = self._amount_int(row.get("cobro", ""))
            comment = "PAGO GC" if amount == charge else "ABONO GC"
            sheet.append([
                folio,
                row.get("uco", ""),
                amount,
                row.get("fecha", ""),
                "Transferencia",
                re.sub(r"\D", "", row.get("fecha", "")),
                comment,
            ])
            folio += 1

        self._fit_columns(sheet)
        return self._workbook_to_base64(workbook)

    def _read_bank_chile(self, content: bytes, filename: str) -> list[ChileBankMovement]:
        rows = self._read_workbook_rows(content, filename)
        header_index = self._find_row_with(rows, ["monto", "descripcionmovimiento", "fecha"])
        headers = [self._normalize_header(value) for value in rows[header_index]]

        def col(name: str) -> int | None:
            clean = self._normalize_header(name)
            return headers.index(clean) if clean in headers else None

        amount_col = col("MONTO")
        desc_col = col("DESCRIPCION MOVIMIENTO")
        date_col = col("FECHA")
        balance_col = col("SALDO")
        document_col = col("N DOCUMENTO")
        branch_col = col("SUCURSAL")
        kind_col = col("CARGO/ABONO")

        if amount_col is None or desc_col is None or date_col is None:
            raise ValueError("La cartola Banco de Chile no tiene las columnas esperadas.")

        movements: list[ChileBankMovement] = []
        for row in rows[header_index + 1 :]:
            amount = self._cell(row, amount_col)
            description = self._cell(row, desc_col)
            date = self._format_date(self._cell(row, date_col))
            if not amount and not description and not date:
                continue

            if kind_col is not None and self._cell(row, kind_col).upper() not in {"", "A"}:
                continue
            if self._amount_int(amount) <= 0:
                continue

            name = self._extract_transfer_name(description)
            if not name:
                continue

            movements.append(
                ChileBankMovement(
                    monto=amount,
                    nombre=name,
                    fecha=date,
                    documento=self._cell(row, document_col),
                    sucursal=self._cell(row, branch_col) or "BC",
                    saldo=self._cell(row, balance_col),
                    rut="",
                )
            )

        if not movements:
            raise ValueError("No se encontraron transferencias positivas en la cartola Banco de Chile.")
        return movements

    def _read_scotiabank(self, content: bytes, filename: str) -> list[ChileBankMovement]:
        rows = self._read_workbook_rows(content, filename)
        header_index = self._find_row_with(rows, ["fecha", "descripcion", "cargos", "abonos"])
        headers = [self._normalize_header(value) for value in rows[header_index]]

        def col(name: str) -> int | None:
            clean = self._normalize_header(name)
            return headers.index(clean) if clean in headers else None

        date_col = col("Fecha")
        desc_col = col("Descripcion")
        debit_col = col("Cargos")
        credit_col = col("Abonos")

        if date_col is None or desc_col is None or credit_col is None:
            raise ValueError("La cartola Scotiabank no tiene las columnas Fecha, Descripcion y Abonos esperadas.")

        movements: list[ChileBankMovement] = []
        for row in rows[header_index + 1 :]:
            date = self._format_date(self._cell(row, date_col))
            description = self._cell(row, desc_col)
            credit = self._cell(row, credit_col)
            debit = self._cell(row, debit_col) if debit_col is not None else ""
            if not date and not description and not credit and not debit:
                continue
            if self._amount_int(credit) <= 0:
                continue

            rut, name = self._extract_scotiabank_transfer(description)
            if not rut and not name:
                continue

            movements.append(
                ChileBankMovement(
                    monto=credit,
                    rut=rut,
                    nombre=name or description,
                    fecha=date,
                    documento=rut,
                    sucursal="Scotiabank",
                    saldo="",
                )
            )

        if not movements:
            raise ValueError("No se encontraron abonos positivos en la cartola Scotiabank.")
        return movements

    def _read_santander(self, content: bytes) -> list[ChileBankMovement]:
        try:
            movements = EdifitoProcessingService()._read_santander_pdf(content)
        except ValueError as exc:
            raise ValueError(str(exc).replace("Santander.", "Santander para Comunidad Feliz."))

        return [
            ChileBankMovement(
                monto=movement.monto,
                rut=movement.rut,
                nombre=movement.nombre,
                fecha=movement.fecha,
                documento=movement.documento,
                sucursal=movement.sucursal,
                saldo="",
            )
            for movement in movements
        ]

    def _read_charges(self, content: bytes) -> list[ComunidadFelizCharge]:
        rows = self._read_workbook_rows(content, "charges.xlsx")
        header_index = self._find_row_with(rows, ["unidad", "residente"])
        headers = [self._normalize_header(value) for value in rows[header_index]]

        def col(*names: str) -> int | None:
            for name in names:
                clean = self._normalize_header(name)
                if clean in headers:
                    return headers.index(clean)
            return None

        uco_col = col("Unidad", "Propiedad")
        rut_col = col("Rut", "RUT", "Documento")
        resident_col = col("Residente", "Copropietario")
        charge_col = col("Total cobrado", "Cobro", "Monto")
        paid_col = col("Abonos", "Pagado", "Pago")
        status_col = col("Estado")
        date_col = col("Fecha ultimo pago", "Fecha Pago")

        if uco_col is None or resident_col is None or charge_col is None:
            raise ValueError("El Excel de boletas no tiene las columnas Unidad, Residente y Total cobrado esperadas.")

        charges: list[ComunidadFelizCharge] = []
        for row in rows[header_index + 1 :]:
            uco = self._cell(row, uco_col)
            resident = self._dedupe_words(self._cell(row, resident_col))
            if not uco:
                continue

            charge = self._cell(row, charge_col) or "0"
            paid = self._cell(row, paid_col) if paid_col is not None else "0"
            paid = paid or "0"
            status = self._cell(row, status_col) if status_col is not None else ""
            if not status:
                status = "Pagado" if self._amount_int(paid) >= self._amount_int(charge) and self._amount_int(charge) > 0 else "Pendiente"

            charges.append(
                ComunidadFelizCharge(
                    uco=uco,
                    rut=self._cell(row, rut_col) if rut_col is not None else "",
                    nombre=resident,
                    cobro=charge,
                    pago=paid,
                    estado=status,
                    fecha_pago=self._format_date(self._cell(row, date_col)) if date_col is not None else "",
                )
            )

        return charges

    def _build_rows(self, movements: list[ChileBankMovement], charges: list[ComunidadFelizCharge]) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        charges_by_rut = {self._normalize_rut(charge.rut): charge for charge in charges if charge.rut}
        for movement in movements:
            normalized_rut = self._normalize_rut(movement.rut)
            best_charge: ComunidadFelizCharge | None = charges_by_rut.get(normalized_rut) if normalized_rut else None
            best_score = 1.0 if best_charge else 0.0

            if not best_charge:
                for charge in charges:
                    score = self._name_similarity(movement.nombre, charge.nombre)
                    if score >= 0.60 and score > best_score:
                        best_score = score
                        best_charge = charge

            row = {
                "monto": movement.monto,
                "rut": movement.rut,
                "nombre": movement.nombre,
                "fecha": movement.fecha,
                "documento": movement.documento,
                "sucursal": movement.sucursal,
                "uco": "",
                "match_rut": "",
                "match_nombre": "",
                "mes": "",
                "cobro": "",
                "pago": "",
                "fecha_pago": "",
                "status": "",
            }

            if best_charge:
                row["uco"] = best_charge.uco
                row["match_rut"] = best_charge.rut if normalized_rut and self._normalize_rut(best_charge.rut) == normalized_rut else ""
                row["match_nombre"] = f"{round(best_score * 100):.0f}% {best_charge.nombre}".strip()
                row["cobro"] = best_charge.cobro
                row["pago"] = best_charge.pago
                row["fecha_pago"] = best_charge.fecha_pago
                row["status"] = self._status_for(movement, best_charge)

            rows.append(row)

        return rows

    def _status_for(self, movement: ChileBankMovement, charge: ComunidadFelizCharge) -> str:
        amount = self._amount_int(movement.monto)
        charge_amount = self._amount_int(charge.cobro)
        paid_amount = self._amount_int(charge.pago)
        status = charge.estado.strip().casefold()

        if status == "pagado" or amount == paid_amount:
            return "Ya procesado"
        if status == "pendiente" and amount <= charge_amount and paid_amount == 0:
            return "Procesar Pago"
        return "Analizar"

    def _summary(self, rows: list[dict[str, str]]) -> dict[str, int]:
        return {
            "total": len(rows),
            "matched": sum(1 for row in rows if row.get("uco")),
            "not_matched": sum(1 for row in rows if not row.get("uco")),
            "procesar_pago": sum(1 for row in rows if row.get("status") == "Procesar Pago"),
            "analizar": sum(1 for row in rows if row.get("status") == "Analizar"),
            "ya_procesado": sum(1 for row in rows if row.get("status") == "Ya procesado"),
        }

    def _read_workbook_rows(self, content: bytes, filename: str) -> list[list[Any]]:
        if filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx"):
            if xlrd is None:
                raise ValueError("Falta dependencia xlrd para leer archivos .xls.")
            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_index(0)
            return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]

        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]

    def _find_row_with(self, rows: list[list[Any]], required: list[str]) -> int:
        required_set = {self._normalize_header(item) for item in required}
        for index, row in enumerate(rows[:80]):
            normalized = {self._normalize_header(value) for value in row}
            if required_set.issubset(normalized):
                return index
        raise ValueError("No se encontro la cabecera esperada en el archivo.")

    def _cell(self, row: list[Any], index: int | None) -> str:
        if index is None or index >= len(row):
            return ""
        value = row[index]
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d/%m/%Y")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _extract_transfer_name(self, description: str) -> str:
        text = re.sub(r"\s+", " ", description or "").strip()
        if not text:
            return ""
        match = re.match(r"^(?:\d+[0-9Kk]?)\s+Transf(?:\.|\s+de)?\s*(.*)$", text, flags=re.IGNORECASE)
        if match:
            return match.group(1).strip()
        match = re.search(r"Transf(?:\.|\s+de)?\s*(.*)$", text, flags=re.IGNORECASE)
        return match.group(1).strip() if match else ""

    def _extract_scotiabank_transfer(self, description: str) -> tuple[str, str]:
        text = re.sub(r"\s+", " ", description or "").strip()
        if not text:
            return "", ""

        match = re.search(r"\b(\d{1,2}\.?\d{3}\.?\d{3}-?[0-9Kk])\b\s*(.*)$", text)
        if not match:
            return "", self._strip_bank_prefix(text)

        rut = match.group(1)
        name = self._strip_bank_prefix(match.group(2).strip())
        return rut, name

    def _strip_bank_prefix(self, value: str) -> str:
        return re.sub(r"^(?:TEF|PROVEEDORE?S?|TRANSFERENCIA|TRASPASO)\s+", "", value or "", flags=re.IGNORECASE).strip()

    def _format_date(self, value: str) -> str:
        text = (value or "").strip()
        if not text:
            return ""
        if re.fullmatch(r"\d+(\.0)?", text):
            try:
                serial = int(float(text))
                return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + serial - 2).strftime("%d/%m/%Y")
            except ValueError:
                return text
        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
            except ValueError:
                continue
        return text

    def _amount_int(self, value: str) -> int:
        clean = re.sub(r"\D", "", value or "")
        return int(clean) if clean else 0

    def _normalize_rut(self, value: str) -> str:
        clean = re.sub(r"[^0-9kK]", "", value or "").upper()
        return clean.lstrip("0")

    def _normalize_header(self, value: Any) -> str:
        text = "" if value is None else str(value)
        text = self._strip_accents(text)
        return re.sub(r"[^a-z0-9]", "", text.casefold())

    def _dedupe_words(self, value: str) -> str:
        seen: set[str] = set()
        words: list[str] = []
        for word in re.split(r"\s+", value or ""):
            clean = word.casefold()
            if clean and clean not in seen:
                seen.add(clean)
                words.append(word)
        return " ".join(words)

    def _name_similarity(self, left: str, right: str, token_threshold: float = 0.84, apellido_boost: float = 1.25) -> float:
        a_tokens = self._normalize_tokens(left)
        b_tokens = self._normalize_tokens(right)
        if not a_tokens or not b_tokens:
            return 0.0

        a_last = set(range(max(0, len(a_tokens) - 2), len(a_tokens)))
        b_last = set(range(max(0, len(b_tokens) - 2), len(b_tokens)))
        used_a: set[int] = set()
        used_b: set[int] = set()
        matches: list[tuple[float, float]] = []

        while True:
            best = token_threshold
            best_pair: tuple[int, int] | None = None
            for i, a in enumerate(a_tokens):
                if i in used_a:
                    continue
                for j, b in enumerate(b_tokens):
                    if j in used_b:
                        continue
                    score = self._jaro_winkler(a, b)
                    if score > best:
                        best = score
                        best_pair = (i, j)
            if best_pair is None:
                break
            i, j = best_pair
            weight = 1.0
            if i in a_last:
                weight *= apellido_boost
            if j in b_last:
                weight *= apellido_boost
            matches.append((best, weight))
            used_a.add(i)
            used_b.add(j)

        weighted = sum(score * weight for score, weight in matches)
        max_a = len(a_tokens) + len(a_last) * (apellido_boost - 1.0)
        max_b = len(b_tokens) + len(b_last) * (apellido_boost - 1.0)
        precision = weighted / max_a if max_a else 0.0
        recall = weighted / max_b if max_b else 0.0
        f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
        if len(matches) >= min(len(a_tokens), len(b_tokens)) and all(score >= 0.90 for score, _ in matches):
            f1 = min(1.0, f1 + 0.05)
        return max(0.0, min(1.0, f1))

    def _normalize_tokens(self, value: str) -> list[str]:
        stop = {"de", "del", "la", "las", "los", "y"}
        clean = self._strip_accents(value).casefold()
        clean = re.sub(r"[^a-z0-9ñ\s]", " ", clean)
        return [token for token in clean.split() if token and token not in stop]

    def _strip_accents(self, value: str) -> str:
        normalized = unicodedata.normalize("NFD", value or "")
        return "".join(char for char in normalized if unicodedata.category(char) != "Mn")

    def _jaro_winkler(self, left: str, right: str, prefix_scale: float = 0.1, max_prefix: int = 4) -> float:
        if left == right:
            return 1.0
        len_left, len_right = len(left), len(right)
        if not len_left or not len_right:
            return 0.0
        match_distance = max(len_left, len_right) // 2 - 1
        left_matches = [False] * len_left
        right_matches = [False] * len_right
        matches = 0

        for i, char in enumerate(left):
            start = max(0, i - match_distance)
            end = min(i + match_distance + 1, len_right)
            for j in range(start, end):
                if right_matches[j] or char != right[j]:
                    continue
                left_matches[i] = right_matches[j] = True
                matches += 1
                break
        if not matches:
            return 0.0

        k = 0
        transpositions = 0
        for i in range(len_left):
            if not left_matches[i]:
                continue
            while not right_matches[k]:
                k += 1
            if left[i] != right[k]:
                transpositions += 1
            k += 1
        m = float(matches)
        jaro = (m / len_left + m / len_right + (m - transpositions / 2.0) / m) / 3.0

        prefix = 0
        for i in range(min(len_left, len_right, max_prefix)):
            if left[i] != right[i]:
                break
            prefix += 1
        return jaro + prefix * prefix_scale * (1 - jaro)

    def _fit_columns(self, sheet: Any) -> None:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 46)

    def _workbook_to_base64(self, workbook: Workbook) -> str:
        output = BytesIO()
        workbook.save(output)
        return base64.b64encode(output.getvalue()).decode("ascii")
