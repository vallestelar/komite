from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
import unicodedata

from openpyxl import load_workbook

from app.models.entities import Unit, UnitContact, User


class EdifitoNeighborsImportService:
    relationship_owner = "copropietario"
    relationship_resident = "residente"

    async def import_assignments(
        self,
        *,
        company_id: str | None,
        condominium_id: str,
        content: bytes,
    ) -> dict[str, Any]:
        rows = self._read_rows(content)
        summary = {
            "rows": len(rows),
            "units_created": 0,
            "units_updated": 0,
            "contacts_created": 0,
            "contacts_updated": 0,
            "contacts_skipped": 0,
            "users_created": 0,
            "users_updated": 0,
            "users_skipped": 0,
        }
        details: list[dict[str, str]] = []

        for row in rows:
            unit, created_unit = await self._upsert_unit(company_id, condominium_id, row)
            summary["units_created" if created_unit else "units_updated"] += 1

            for relationship, prefix in (
                (self.relationship_owner, "owner"),
                (self.relationship_resident, "resident"),
            ):
                full_name = self._clean(row.get(f"{prefix}_name"))
                if not full_name:
                    summary["contacts_skipped"] += 1
                    continue

                user_status = await self._classify_user(
                    company_id=company_id,
                    email=self._clean_email(row.get(f"{prefix}_email")),
                )
                contact, created_contact = await self._upsert_contact(
                    company_id=company_id,
                    condominium_id=condominium_id,
                    unit=unit,
                    relationship_type=relationship,
                    full_name=full_name,
                    email=self._clean_email(row.get(f"{prefix}_email")),
                    phone=self._clean(row.get(f"{prefix}_phone")),
                    document_number=self._clean(row.get(f"{prefix}_rut")),
                    address=self._clean(row.get(f"{prefix}_address")),
                )
                summary["contacts_created" if created_contact else "contacts_updated"] += 1
                summary[f"users_{user_status}"] += 1
                details.append(
                    {
                        "unit": unit.identifier,
                        "relationship_type": contact.relationship_type,
                        "full_name": contact.full_name,
                        "status": "creado" if created_contact else "actualizado",
                    }
                )

        return {"summary": summary, "items": details[:200]}

    async def preview_assignments(
        self,
        *,
        company_id: str | None,
        condominium_id: str,
        content: bytes,
    ) -> dict[str, Any]:
        rows = self._read_rows(content)
        summary = {
            "rows": len(rows),
            "units_created": 0,
            "units_updated": 0,
            "contacts_created": 0,
            "contacts_updated": 0,
            "contacts_skipped": 0,
            "users_created": 0,
            "users_updated": 0,
            "users_skipped": 0,
        }
        details: list[dict[str, str]] = []

        for row in rows:
            identifier = self._clean(row["uco"]) or "Sin identificador"
            unit = await Unit.get_or_none(condominium_id=condominium_id, identifier=identifier)
            summary["units_updated" if unit else "units_created"] += 1

            for relationship, prefix in (
                (self.relationship_owner, "owner"),
                (self.relationship_resident, "resident"),
            ):
                full_name = self._clean(row.get(f"{prefix}_name"))
                if not full_name:
                    summary["contacts_skipped"] += 1
                    continue

                email = self._clean_email(row.get(f"{prefix}_email"))
                document_number = self._clean(row.get(f"{prefix}_rut"))
                existing_contact = await self._find_contact(unit, relationship, email, document_number, full_name) if unit else None
                contact_status = "updated" if existing_contact else "created"
                summary[f"contacts_{contact_status}"] += 1

                user_status = await self._classify_user(company_id=company_id, email=email)
                summary[f"users_{user_status}"] += 1
                details.append(
                    {
                        "unit": identifier,
                        "relationship_type": relationship,
                        "full_name": full_name,
                        "status": "se va a crear" if contact_status == "created" else "se va a actualizar",
                    }
                )

        return {"summary": summary, "items": details[:200]}

    def _read_rows(self, content: bytes) -> list[dict[str, Any]]:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = workbook["Datos"] if "Datos" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        header_row = self._find_header_row(sheet)
        headers = [self._cell_text(cell.value) for cell in sheet[header_row]]
        indexes = self._build_indexes(headers)

        required = ["uco", "tipo_unidad"]
        if any(indexes.get(name) is None for name in required):
            raise ValueError("El informe no tiene las columnas Uco y Tipo Unidad esperadas.")

        rows: list[dict[str, Any]] = []
        for raw in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            uco = self._value(raw, indexes["uco"])
            if not self._clean(uco):
                continue

            rows.append(
                {
                    "uco": self._cell_text(uco),
                    "proration_total": self._value(raw, indexes.get("prorrateo_total")),
                    "unit_type": self._cell_text(self._value(raw, indexes.get("tipo_unidad"))),
                    "allocation_number": self._value(raw, indexes.get("numero_asignacion")),
                    "allocation_identifier": self._cell_text(self._value(raw, indexes.get("asignaciones"))),
                    "proration": self._value(raw, indexes.get("prorrateo")),
                    "assignment_date": self._value(raw, indexes.get("fecha")),
                    "owner_rut": self._cell_text(self._value(raw, indexes.get("owner_rut"))),
                    "owner_name": self._cell_text(self._value(raw, indexes.get("owner_name"))),
                    "owner_email": self._cell_text(self._value(raw, indexes.get("owner_email"))),
                    "owner_address": self._cell_text(self._value(raw, indexes.get("owner_address"))),
                    "owner_phone": self._cell_text(self._value(raw, indexes.get("owner_phone"))),
                    "resident_rut": self._cell_text(self._value(raw, indexes.get("resident_rut"))),
                    "resident_name": self._cell_text(self._value(raw, indexes.get("resident_name"))),
                    "resident_email": self._cell_text(self._value(raw, indexes.get("resident_email"))),
                    "resident_address": self._cell_text(self._value(raw, indexes.get("resident_address"))),
                    "resident_phone": self._cell_text(self._value(raw, indexes.get("resident_phone"))),
                }
            )

        if not rows:
            raise ValueError("El informe de asignaciones no contiene filas para importar.")
        return rows

    async def _upsert_unit(self, company_id: str | None, condominium_id: str, row: dict[str, Any]) -> tuple[Unit, bool]:
        identifier = self._clean(row["uco"]) or "Sin identificador"
        unit = await Unit.get_or_none(condominium_id=condominium_id, identifier=identifier)
        created = unit is None
        if not unit:
            unit = Unit(company_id=company_id, condominium_id=condominium_id, identifier=identifier)

        unit.company_id = company_id
        unit.external_code = identifier
        unit.unit_type = self._map_unit_type(row.get("unit_type"))
        unit.allocation_number = self._int_or_none(row.get("allocation_number"))
        unit.allocation_identifier = self._clean(row.get("allocation_identifier"))
        unit.proration_total = self._decimal_or_none(row.get("proration_total"))
        unit.proration = self._decimal_or_none(row.get("proration"))
        unit.assignment_date = self._date_or_none(row.get("assignment_date"))
        await unit.save()
        return unit, created

    async def _upsert_contact(
        self,
        *,
        company_id: str | None,
        condominium_id: str,
        unit: Unit,
        relationship_type: str,
        full_name: str,
        email: str | None,
        phone: str | None,
        document_number: str | None,
        address: str | None,
    ) -> tuple[UnitContact, bool]:
        contact = await self._find_contact(unit, relationship_type, email, document_number, full_name)
        created = contact is None
        user = await self._sync_user(company_id, email, full_name, phone, document_number, address)

        if not contact:
            contact = UnitContact(
                company_id=company_id,
                condominium_id=condominium_id,
                unit=unit,
                relationship_type=relationship_type,
                full_name=full_name,
            )

        contact.company_id = company_id
        contact.condominium_id = condominium_id
        contact.unit = unit
        contact.user = user
        contact.full_name = full_name
        contact.email = email
        contact.phone = phone
        contact.document_type = "rut" if document_number else None
        contact.document_number = document_number
        contact.address = address
        contact.status = "active"
        await contact.save()
        return contact, created

    async def _find_contact(
        self,
        unit: Unit,
        relationship_type: str,
        email: str | None,
        document_number: str | None,
        full_name: str,
    ) -> UnitContact | None:
        base = UnitContact.filter(unit=unit, relationship_type=relationship_type)
        if email:
            return await base.filter(email=email).first()
        if document_number:
            return await base.filter(document_number=document_number).first()
        return await base.filter(full_name=full_name).first()

    async def _classify_user(self, company_id: str | None, email: str | None) -> str:
        if not email:
            return "skipped"
        user = await User.get_or_none(email=email)
        if not user:
            return "created"
        if company_id and str(user.company_id) == str(company_id):
            return "updated"
        return "skipped"

    async def _sync_user(
        self,
        company_id: str | None,
        email: str | None,
        full_name: str,
        phone: str | None,
        document_number: str | None,
        address: str | None,
    ) -> User | None:
        if not email:
            return None
        user = await User.get_or_none(email=email)
        if not user:
            return await User.create(
                company_id=company_id,
                email=email,
                full_name=full_name,
                phone=phone,
                document_type="rut" if document_number else None,
                document_number=document_number,
                address=address,
                status="active",
            )
        if company_id and str(user.company_id) == str(company_id):
            user.full_name = full_name
            user.phone = phone
            user.document_type = "rut" if document_number else user.document_type
            user.document_number = document_number
            user.address = address
            await user.save()
            return user
        return None

    def _find_header_row(self, sheet) -> int:
        for index, row in enumerate(sheet.iter_rows(), start=1):
            normalized = {self._normalize_header(self._cell_text(cell.value)) for cell in row}
            if {"uco", "tipounidad", "copropietario"}.issubset(normalized):
                return index
        raise ValueError("No se encontr\u00f3 la cabecera del informe de asignaciones.")

    def _build_indexes(self, headers: list[str]) -> dict[str, int | None]:
        groups: dict[str, list[int]] = {}
        for index, header in enumerate(headers):
            normalized = self._normalize_header(header)
            groups.setdefault(normalized, []).append(index)

        def idx(name: str, occurrence: int = 0) -> int | None:
            values = groups.get(self._normalize_header(name), [])
            return values[occurrence] if len(values) > occurrence else None

        return {
            "uco": idx("Uco"),
            "prorrateo_total": idx("% Prorrateo total"),
            "tipo_unidad": idx("Tipo Unidad"),
            "numero_asignacion": idx("N Asign."),
            "asignaciones": idx("Asignaciones"),
            "prorrateo": idx("% Prorrateo"),
            "fecha": idx("Fecha"),
            "owner_rut": idx("Rut", 0),
            "owner_name": idx("Copropietario"),
            "owner_email": idx("Email", 0),
            "owner_address": idx("Direcci\u00f3n", 0),
            "owner_phone": idx("Tel\u00e9fono", 0),
            "resident_rut": idx("Rut", 1),
            "resident_name": idx("Residente"),
            "resident_email": idx("Email", 1),
            "resident_address": idx("Direcci\u00f3n", 1),
            "resident_phone": idx("Tel\u00e9fono", 1),
        }

    def _value(self, row: tuple[Any, ...], index: int | None) -> Any:
        if index is None or index >= len(row):
            return None
        return row[index]

    def _cell_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _normalize_header(self, value: str) -> str:
        text = unicodedata.normalize("NFKD", value)
        return "".join(ch for ch in text.casefold() if ch.isalnum() and not unicodedata.combining(ch))

    def _clean(self, value: Any) -> str | None:
        text = self._cell_text(value)
        return text if text else None

    def _clean_email(self, value: Any) -> str | None:
        text = self._clean(value)
        return text.casefold() if text and "@" in text else None

    def _map_unit_type(self, value: Any) -> str:
        text = (self._clean(value) or "").casefold()
        if "departamento" in text:
            return "apartment"
        if "casa" in text:
            return "house"
        if "bodega" in text:
            return "storage"
        if "estacionamiento" in text:
            return "parking"
        if "oficina" in text:
            return "office"
        return "other" if text else "apartment"

    def _int_or_none(self, value: Any) -> int | None:
        text = self._clean(value)
        if not text:
            return None
        try:
            return int(float(text.replace(",", ".")))
        except ValueError:
            return None

    def _decimal_or_none(self, value: Any) -> Decimal | None:
        text = self._clean(value)
        if not text:
            return None
        try:
            return Decimal(text.replace(",", "."))
        except (InvalidOperation, ValueError):
            return None

    def _date_or_none(self, value: Any) -> date | None:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = self._clean(value)
        if not text:
            return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None
