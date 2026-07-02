from __future__ import annotations

import base64
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader


@dataclass
class BankMovement:
    monto: str
    rut: str
    nombre: str
    fecha: str
    documento: str
    sucursal: str


@dataclass
class Assignment:
    uco: str
    rut_copropietario: str
    copropietario: str
    rut_residente: str
    residente: str


@dataclass
class ChargeDetail:
    uco: str
    mes: str
    cobro: str
    descuento: str
    pago: str
    fecha_pago: str
    status: str | None = None


class EdifitoProcessingService:
    output_columns = [
        ("monto", "Monto"),
        ("rut", "RUT"),
        ("nombre", "Nombre"),
        ("fecha", "Fecha"),
        ("documento", "Documento"),
        ("sucursal", "Sucursal"),
        ("uco", "UCO"),
        ("match_rut", "MatchRUT"),
        ("match_nombre", "MatchNombre"),
        ("mes", "Mes"),
        ("cobro", "Cobro"),
        ("pago", "Pago"),
        ("fecha_pago", "FechaPago"),
        ("status", "Status"),
    ]

    def process(
        self,
        *,
        bank_name: str,
        bank_statement: bytes,
        assignments_file: bytes,
        charge_detail_file: bytes,
    ) -> dict[str, Any]:
        if "santander" not in (bank_name or "").casefold():
            raise ValueError("Por ahora Edifito solo procesa cartolas Santander.")

        movements = self._read_santander_pdf(bank_statement)
        assignments = self._read_assignments(assignments_file)
        charge_details = self._read_charge_details(charge_detail_file)

        rows = self._build_rows(movements, assignments, charge_details)
        return {
            "rows": rows,
            "summary": self._summary(rows),
            "download_filename": f"Movimientos_{datetime.now().strftime('%d_%m_%Y_%H%M')}.xlsx",
            "download_base64": self._build_workbook_base64(rows),
        }

    def _read_santander_pdf(self, content: bytes) -> list[BankMovement]:
        reader = PdfReader(BytesIO(content))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        if "santander" not in text.casefold() and "consulta de movimientos" not in text.casefold():
            raise ValueError("La cartola no parece corresponder al formato Santander esperado.")

        lines = [line.strip() for line in re.split(r"\r\n|\r|\n", text) if line.strip()]
        logical_records: list[str] = []
        current: list[str] = []

        for line in lines:
            if re.match(r"^\$ ?-?\d", line):
                if current:
                    logical_records.append(" ".join(current).strip())
                    current = []
            current.append(line)

        if current:
            logical_records.append(" ".join(current).strip())

        pattern = re.compile(r"\$ ?([-0-9.]+)\s+(.*?)\s+(\d{2}/\d{2}/\d{4})\s+\$ ?([-0-9.]+)\s+(\w+)\s+(.*)")
        movements: list[BankMovement] = []

        for record in logical_records:
            match = pattern.search(record)
            if not match:
                continue

            amount = match.group(1).strip()
            if amount.startswith("-"):
                continue

            description = match.group(2).strip()
            rut = ""
            name = description
            rut_match = re.match(r"^([\dKk]+)\s+(.*)", description)
            if rut_match:
                rut = rut_match.group(1).lstrip("0").upper()
                name = re.sub(r"^Transf(?:\.|\s+de:?\.?)\s+", "", rut_match.group(2), flags=re.IGNORECASE).strip()

            name = self._clean_text(name).upper()
            if name == "KHIPU SPA":
                continue

            movements.append(
                BankMovement(
                    monto=amount,
                    rut=rut,
                    nombre=name,
                    fecha=match.group(3).strip(),
                    documento=match.group(5).strip(),
                    sucursal=self._clean_sucursal(match.group(6).strip()),
                )
            )

        if not movements:
            raise ValueError("No se encontraron movimientos positivos en la cartola Santander.")

        return movements

    def _read_assignments(self, content: bytes) -> list[Assignment]:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = workbook["Datos"] if "Datos" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        header_row = self._find_header_row(sheet, "Uco", "Rut", "Copropietario")
        headers = [self._cell_text(cell.value) for cell in sheet[header_row]]
        header_indexes: dict[str, list[int]] = {}
        for idx, value in enumerate(headers):
            clean = self._normalize_header(value)
            if clean:
                header_indexes.setdefault(clean, []).append(idx)

        def idx(*names: str, occurrence: int = 0) -> int | None:
            for name in names:
                clean = self._normalize_header(name)
                matches = header_indexes.get(clean, [])
                if len(matches) > occurrence:
                    return matches[occurrence]
            return None

        uco_idx = idx("Uco")
        rut_owner_idx = idx("Rut", occurrence=0)
        owner_idx = idx("Copropietario")
        rut_resident_idx = idx("Rut", occurrence=1)
        resident_idx = idx("Residente")

        if uco_idx is None or rut_owner_idx is None or owner_idx is None:
            raise ValueError("El informe de asignaciones no tiene las columnas Uco, Rut y Copropietario esperadas.")

        assignments: list[Assignment] = []
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            uco = self._cell_text(row[uco_idx] if uco_idx < len(row) else "")
            if not uco:
                break

            assignments.append(
                Assignment(
                    uco=uco,
                    rut_copropietario=self._cell_text(row[rut_owner_idx] if rut_owner_idx < len(row) else ""),
                    copropietario=self._cell_text(row[owner_idx] if owner_idx < len(row) else ""),
                    rut_residente=self._cell_text(row[rut_resident_idx] if rut_resident_idx is not None and rut_resident_idx < len(row) else ""),
                    residente=self._cell_text(row[resident_idx] if resident_idx is not None and resident_idx < len(row) else ""),
                )
            )

        return assignments

    def _read_charge_details(self, content: bytes) -> list[ChargeDetail]:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        sheet = workbook["Datos"] if "Datos" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))

        field_row_idx = None
        for index, row in enumerate(rows):
            normalized = [self._normalize_header(self._cell_text(value)) for value in row]
            if "cobro" in normalized and "pago" in normalized and any(value == "fechapago" for value in normalized):
                field_row_idx = index
                break

        if field_row_idx is None or field_row_idx == 0:
            raise ValueError("El informe en detalle no tiene las columnas Cobro, Pago y Fecha Pago esperadas.")

        month_row = rows[field_row_idx - 1]
        details: list[ChargeDetail] = []
        for row in rows[field_row_idx + 1 :]:
            uco = self._cell_text(row[0] if len(row) else "")
            if not uco:
                continue

            for col in range(1, len(row), 4):
                month = self._cell_text(month_row[col] if col < len(month_row) else "")
                if not month:
                    continue

                details.append(
                    ChargeDetail(
                        uco=uco,
                        mes=month,
                        cobro=self._cell_text(row[col] if col < len(row) else "0") or "0",
                        descuento=self._cell_text(row[col + 1] if col + 1 < len(row) else "0") or "0",
                        pago=self._cell_text(row[col + 2] if col + 2 < len(row) else "0") or "0",
                        fecha_pago=self._format_date(row[col + 3] if col + 3 < len(row) else "-"),
                    )
                )

        return details

    def _build_rows(
        self,
        movements: list[BankMovement],
        assignments: list[Assignment],
        charge_details: list[ChargeDetail],
    ) -> list[dict[str, str]]:
        assignment_by_owner = {self._normalize_rut(item.rut_copropietario): item for item in assignments if item.rut_copropietario}
        assignment_by_resident = {self._normalize_rut(item.rut_residente): item for item in assignments if item.rut_residente}

        rows: list[dict[str, str]] = []
        for movement in movements:
            rut = self._normalize_rut(movement.rut)
            assignment = assignment_by_owner.get(rut)
            match_rut = ""
            match_name = ""

            if assignment:
                match_rut = assignment.rut_copropietario
                match_name = assignment.copropietario.upper()
            else:
                assignment = assignment_by_resident.get(rut)
                if assignment:
                    match_rut = assignment.rut_residente
                    match_name = assignment.residente.upper()

            row = {
                "monto": movement.monto,
                "rut": movement.rut,
                "nombre": movement.nombre,
                "fecha": movement.fecha,
                "documento": movement.documento,
                "sucursal": movement.sucursal,
                "uco": assignment.uco if assignment else "",
                "match_rut": match_rut,
                "match_nombre": match_name,
                "mes": "",
                "cobro": "",
                "pago": "",
                "fecha_pago": "",
                "status": "",
            }
            rows.append(row)

        for row in rows:
            detail = self._select_charge_detail(charge_details, row)
            if not detail:
                continue
            row.update(
                {
                    "mes": detail.mes,
                    "cobro": detail.cobro,
                    "pago": detail.pago,
                    "fecha_pago": detail.fecha_pago,
                    "status": detail.status or "",
                }
            )

        return rows

    def _select_charge_detail(self, details: list[ChargeDetail], movement: dict[str, str]) -> ChargeDetail | None:
        if not movement.get("uco"):
            return None

        uco_details = [
            detail
            for detail in details
            if self._normalize_uco(detail.uco) == self._normalize_uco(movement["uco"])
        ]
        if not uco_details:
            return None

        selected = next(
            (
                detail
                for detail in uco_details
                if self._normalize_value(detail.pago) == self._normalize_value(movement["monto"])
                and self._normalize_date(detail.fecha_pago) == self._normalize_date(movement["fecha"])
            ),
            None,
        )
        if selected:
            selected.status = "Ya procesado"
            return selected

        selected = next(
            (
                detail
                for detail in reversed(uco_details)
                if self._normalize_value(detail.cobro) == self._normalize_value(movement["monto"])
            ),
            None,
        )
        if selected:
            selected.status = "Procesar Pago"
            return selected

        for index in range(len(uco_details) - 1, 0, -1):
            current = uco_details[index]
            previous = uco_details[index - 1]
            current_pending = self._normalize_value(current.pago) in {"", "0"} and current.fecha_pago.strip() == "-"
            previous_paid = self._normalize_value(previous.pago) not in {"", "0"} and previous.fecha_pago.strip() != "-"
            if current_pending and previous_paid:
                current.status = "Analizar"
                return current

        selected = uco_details[-1]
        selected.status = "Analizar"
        return selected

    def _build_workbook_base64(self, rows: list[dict[str, str]]) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Movimientos"
        headers = [label for _, label in self.output_columns]
        sheet.append(headers)

        header_fill = PatternFill("solid", fgColor="D9E2F3")
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        for row in rows:
            sheet.append([row.get(key, "") for key, _ in self.output_columns])

        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 42)

        output = BytesIO()
        workbook.save(output)
        return base64.b64encode(output.getvalue()).decode("ascii")

    def _summary(self, rows: list[dict[str, str]]) -> dict[str, int]:
        return {
            "total": len(rows),
            "matched": sum(1 for row in rows if row.get("uco")),
            "not_matched": sum(1 for row in rows if not row.get("uco")),
            "procesar_pago": sum(1 for row in rows if row.get("status") == "Procesar Pago"),
            "analizar": sum(1 for row in rows if row.get("status") == "Analizar"),
            "ya_procesado": sum(1 for row in rows if row.get("status") == "Ya procesado"),
        }

    def _find_header_row(self, sheet: Any, *required: str) -> int:
        required_headers = {self._normalize_header(value) for value in required}
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 40)):
            row_headers = {self._normalize_header(self._cell_text(cell.value)) for cell in row}
            if required_headers.issubset(row_headers):
                return row[0].row
        raise ValueError("No se encontro la cabecera esperada en el informe de asignaciones.")

    def _cell_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%d-%m-%Y")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _format_date(self, value: Any) -> str:
        text = self._cell_text(value)
        if not text:
            return "-"
        return self._normalize_date(text)

    def _normalize_header(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]", "", value.casefold())

    def _normalize_rut(self, value: str) -> str:
        clean = re.sub(r"[^0-9kK]", "", value or "").upper()
        return clean.lstrip("0")

    def _normalize_uco(self, value: str) -> str:
        return (value or "").strip().upper()

    def _normalize_value(self, value: str) -> str:
        clean = re.sub(r"\D", "", value or "")
        return clean.lstrip("0") or "0"

    def _normalize_date(self, value: str) -> str:
        text = (value or "").strip()
        if not text or text == "-":
            return "-"

        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, fmt).strftime("%d-%m-%Y")
            except ValueError:
                continue
        return text

    def _clean_text(self, value: str) -> str:
        return re.sub(r"\s+", " ", value or "").strip()

    def _clean_sucursal(self, value: str) -> str:
        index = value.casefold().find("consulta")
        if index >= 0:
            value = value[:index]
        return self._clean_text(value)
